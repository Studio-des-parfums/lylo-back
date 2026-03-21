"""Service de génération de formules de parfum.

Charge les données du coffret (XLSX) une seule fois en mémoire,
puis génère 2 formules personnalisées à partir des réponses utilisateur.
Calcule les quantités en ml pour 3 formats (10ml, 30ml, 50ml).

Architecture :
  1. Réponses questionnaire → scoring direct des notes (note_scoring_mapping.json)
  2. Sélection des meilleures notes par catégorie (top / heart / base), indépendamment
  3. Profil déduit des notes sélectionnées (profile_1 / profile_2 du XLSX)
  4. Type de formule déduit du genre du profil (frais / mix / puissant)
  5. Calcul des ml selon le tableau des formules
"""

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

from app.data.choice_profile_mapping import (
    INGREDIENT_EN_TO_FR,
    PROFILE_DESCRIPTIONS,
    PROFILE_DESCRIPTIONS_EN,
    PROFILE_GENDERS,
)
from app.data.questions import EN_TO_FR_CHOICES
from app.config import get_settings
from app.services import mail_service, session_store

# ── Chemins ───────────────────────────────────────────────────────────
_DATA_DIR = Path(__file__).resolve().parent.parent / "data"
_XLSX_PATH = _DATA_DIR / "Coffret-description.xlsx"
_NOTE_SCORING_PATH = _DATA_DIR / "note_scoring_mapping.json"

# ── Cache mémoire (chargé une seule fois) ─────────────────────────────
_coffret: dict | None = None
_note_scoring_mapping: dict | None = None


# ── Normalisation des noms de profils ─────────────────────────────────
_PROFILE_NORMALIZE = {
    "stratégist": "Strategist",
    "strategist": "Strategist",
    "disrupteur": "Disruptor",
    "disruptor": "Disruptor",
    "trail blazer": "Trailblazer",
    "trailblazer": "Trailblazer",
    "visionary": "Visionary",
    "visionnary": "Visionary",
    "innovator": "Innovator",
    "creator": "Creator",
    "influencer": "Influencer",
    "icon": "Icon",
    "cosy": "Cosy",
}


def _normalize_profile(raw: str | None) -> str | None:
    if not raw:
        return None
    return _PROFILE_NORMALIZE.get(raw.strip().lower(), raw.strip())


# ── Chargement du XLSX ────────────────────────────────────────────────

def _load_coffret() -> dict:
    """Parse le XLSX et retourne les ingrédients + allergènes."""
    wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)

    ws = wb[wb.sheetnames[0]]
    ingredients = []

    note_ranges = [
        ("top", 7, 16),
        ("heart", 21, 30),
        ("base", 35, 44),
    ]

    for note_type, start, end in note_ranges:
        for row in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=8):
            position = row[0].value
            if not position:
                continue
            ingredients.append({
                "position": position,
                "name": row[1].value,
                "family": row[2].value,
                "description": row[3].value,
                "note_type": note_type,
                "profile_1": _normalize_profile(row[6].value),
                "profile_2": _normalize_profile(row[7].value),
            })

    ws2 = wb["ALLERGENS"]

    allergen_blocks = [
        (4, 6, 31),
        (33, 35, 60),
        (62, 64, 89),
    ]

    allergen_map: dict[str, set[str]] = defaultdict(set)

    for header_row, data_start, data_end in allergen_blocks:
        header_cells = list(ws2.iter_rows(
            min_row=header_row, max_row=header_row,
            min_col=2, max_col=11,
        ))[0]
        col_to_ingredient = {}
        for cell in header_cells:
            if cell.value:
                col_to_ingredient[cell.column] = cell.value.strip()

        for row in ws2.iter_rows(
            min_row=data_start, max_row=data_end,
            min_col=1, max_col=11,
        ):
            allergen_name = row[0].value
            if not allergen_name:
                continue
            allergen_name = allergen_name.strip()
            for cell in row[1:]:
                if cell.value and str(cell.value).strip().lower() == "x":
                    ingredient_name = col_to_ingredient.get(cell.column)
                    if ingredient_name:
                        allergen_map[ingredient_name].add(allergen_name)

    wb.close()

    return {
        "ingredients": ingredients,
        "allergen_map": dict(allergen_map),
    }


def _get_coffret() -> dict:
    global _coffret
    if _coffret is None:
        _coffret = _load_coffret()
    return _coffret


def _get_note_scoring_mapping() -> dict:
    global _note_scoring_mapping
    if _note_scoring_mapping is None:
        with open(_NOTE_SCORING_PATH, encoding="utf-8") as f:
            _note_scoring_mapping = json.load(f)
    return _note_scoring_mapping


# ── Configuration des types de formules ──────────────────────────────
#
# Tableau des formules :
#
# Frais/léger  (3T + 3H + 2B)
#   10ml : top=1ml  heart=1ml  base=2ml  booster=1ml   → total=10ml  (+1ml booster=11ml)
#   30ml : top=3ml  heart=3ml  base=6ml  booster=3ml   → total=30ml  (+3ml booster=33ml)
#   50ml : top=5ml  heart=5ml  base=10ml booster=5ml   → total=50ml  (+5ml booster=55ml)
#
# Mix          (2T + 3H + 2B)  — le booster "remplace" la note de tête manquante
#   10ml : top=1ml  heart=1ml  base=2ml  booster=1ml   → notes=9ml   total=10ml
#   30ml : top=3ml  heart=3ml  base=6ml  booster=3ml   → notes=27ml  total=30ml
#   50ml : top=5ml  heart=5ml  base=10ml booster=5ml   → notes=45ml  total=50ml
#
# Puissant/fort (2T + 2H + 3B)
#   10ml : top=1ml  heart=1ml  base=2ml  booster=1ml   → total=10ml  (+1ml booster=11ml)
#   30ml : top=2ml  heart=4ml  base=6ml  booster=3ml   → total=30ml  (+3ml booster=33ml)
#   50ml : top=4ml  heart=6ml  base=10ml booster=5ml   → total=50ml  (+5ml booster=55ml)

_FORMULA_TYPE_CONFIGS: dict[str, dict] = {
    "frais": {
        "note_counts": {"top": 3, "heart": 3, "base": 2},
        "sizes": {
            10: {"top_ml": 1, "heart_ml": 1, "base_ml": 2, "booster_ml": 1},
            30: {"top_ml": 3, "heart_ml": 3, "base_ml": 6, "booster_ml": 3},
            50: {"top_ml": 5, "heart_ml": 5, "base_ml": 10, "booster_ml": 5},
        },
    },
    "mix": {
        "note_counts": {"top": 2, "heart": 3, "base": 2},
        "sizes": {
            10: {"top_ml": 1, "heart_ml": 1, "base_ml": 2, "booster_ml": 1},
            30: {"top_ml": 3, "heart_ml": 3, "base_ml": 6, "booster_ml": 3},
            50: {"top_ml": 5, "heart_ml": 5, "base_ml": 10, "booster_ml": 5},
        },
    },
    "puissant": {
        "note_counts": {"top": 2, "heart": 2, "base": 3},
        "sizes": {
            10: {"top_ml": 1, "heart_ml": 1, "base_ml": 2, "booster_ml": 1},
            30: {"top_ml": 2, "heart_ml": 4, "base_ml": 6, "booster_ml": 3},
            50: {"top_ml": 4, "heart_ml": 6, "base_ml": 10, "booster_ml": 5},
        },
    },
}

# Genre du profil → type de formule
_PROFILE_GENDER_TO_FORMULA_TYPE: dict[str, str] = {
    "masculine": "puissant",
    "feminine": "frais",
    "unisex": "mix",
}


# ── Scoring des notes ─────────────────────────────────────────────────

def _resolve_en_choice(choice: str, qid: int, q_choices_mapping: dict) -> str:
    """Résout un choix de réponse (FR ou EN) vers la clé anglaise du mapping."""
    # Correspondance directe
    if choice in q_choices_mapping:
        return choice

    # Correspondance par préfixe (avant " - ")
    choice_prefix = choice.split(" - ")[0].strip().lower()
    for key in q_choices_mapping:
        if key.split(" - ")[0].strip().lower() == choice_prefix:
            return key

    # Reverse lookup FR → EN via EN_TO_FR_CHOICES
    en_to_fr = EN_TO_FR_CHOICES.get(qid, {})
    for en_key, fr_val in en_to_fr.items():
        if fr_val.lower() == choice.lower():
            en_prefix = en_key.split(" - ")[0].strip().lower()
            for mapping_key in q_choices_mapping:
                if mapping_key.split(" - ")[0].strip().lower() == en_prefix:
                    return mapping_key
            return en_key

    return choice


def _score_notes(answers: dict) -> dict[str, dict[str, float]]:
    """Score toutes les notes du coffret à partir des réponses au questionnaire.

    Scoring par catégorie indépendant (top / heart / base).
    top_2 choice → poids ×2, bottom_2 choice → poids ×−1.
    Fallback familles pour les notes sans score direct.
    """
    note_mapping = _get_note_scoring_mapping()
    coffret = _get_coffret()

    # Index des noms de notes par catégorie (noms anglais du coffret)
    note_by_category: dict[str, set[str]] = {"top": set(), "heart": set(), "base": set()}
    for ing in coffret["ingredients"]:
        note_by_category[ing["note_type"]].add(ing["name"])

    scores: dict[str, dict[str, float]] = {
        "top": defaultdict(float),
        "heart": defaultdict(float),
        "base": defaultdict(float),
    }

    for qid_str, answer_data in answers.items():
        qid = int(qid_str)
        q_choices = note_mapping["questions"].get(str(qid), {}).get("choices", {})

        if isinstance(answer_data, str):
            answer_data = json.loads(answer_data)

        def apply_choice(choice: str, weight: float) -> None:
            en_choice = _resolve_en_choice(choice, qid, q_choices)
            choice_data = q_choices.get(en_choice, {})

            # Scores directs sur les notes nommées
            note_scores = choice_data.get("notes", {})
            for note_name, score in note_scores.items():
                for cat, cat_notes in note_by_category.items():
                    if note_name in cat_notes:
                        scores[cat][note_name] += score * weight

            # Fallback familles pour les notes non citées directement
            families = choice_data.get("families", {})
            if families:
                for ing in coffret["ingredients"]:
                    if ing["name"] in note_scores:
                        continue
                    fam_score = families.get(ing["family"], 0)
                    if fam_score:
                        scores[ing["note_type"]][ing["name"]] += fam_score * weight

        for choice in answer_data.get("top_2", []):
            apply_choice(choice, weight=2.0)
        for choice in answer_data.get("bottom_2", []):
            apply_choice(choice, weight=-1.0)

    return {cat: dict(s) for cat, s in scores.items()}


# ── Sélection des notes ───────────────────────────────────────────────

def _select_notes_by_score(
    note_scores: dict[str, dict[str, float]],
    max_per_cat: int = 3,
    excluded_names: set[str] | None = None,
    blocked_names: set[str] | None = None,
) -> dict[str, list[dict]]:
    """Sélectionne les meilleures notes par catégorie selon les scores.

    Si tous les scores sont nuls, utilise l'ordre de position comme fallback.
    """
    coffret = _get_coffret()
    excluded = excluded_names or set()
    blocked = blocked_names or set()

    cat_to_key = {"top": "top_notes", "heart": "heart_notes", "base": "base_notes"}
    result: dict[str, list[dict]] = {}

    for cat, key in cat_to_key.items():
        cat_scores = note_scores.get(cat, {})
        cat_ingredients = [
            ing for ing in coffret["ingredients"]
            if ing["note_type"] == cat
            and ing["name"] not in excluded
            and ing["name"] not in blocked
        ]

        # Tri : score décroissant, puis position alphabétique comme tiebreaker
        sorted_ing = sorted(
            cat_ingredients,
            key=lambda x: (-cat_scores.get(x["name"], 0), x["position"]),
        )
        selected = sorted_ing[:max_per_cat]

        result[key] = [
            {
                "position": ing["position"],
                "name": ing["name"],
                "family": ing["family"],
                "description": ing["description"],
            }
            for ing in selected
        ]

    return result


def _derive_profile_from_notes(selected_notes: dict[str, list[dict]]) -> str:
    """Dérive le profil dominant à partir des notes sélectionnées.

    Utilise les champs profile_1 (poids 2) et profile_2 (poids 1) du XLSX.
    """
    coffret = _get_coffret()
    ing_by_name = {ing["name"]: ing for ing in coffret["ingredients"]}

    profile_counts: dict[str, int] = defaultdict(int)
    for key in ("top_notes", "heart_notes", "base_notes"):
        for note in selected_notes.get(key, []):
            ing = ing_by_name.get(note["name"])
            if ing:
                if ing["profile_1"]:
                    profile_counts[ing["profile_1"]] += 2
                if ing["profile_2"]:
                    profile_counts[ing["profile_2"]] += 1

    if not profile_counts:
        return "Trailblazer"

    return max(profile_counts, key=lambda p: profile_counts[p])


def _classify_formula_type(profile_name: str) -> str:
    """Classifie le type de formule selon le genre du profil dominant."""
    profile_gender = PROFILE_GENDERS.get(profile_name, "unisex")
    return _PROFILE_GENDER_TO_FORMULA_TYPE[profile_gender]


def _get_blocked_ingredients(user_allergens: list[str] | None) -> set[str]:
    """Retourne les noms d'ingrédients bloqués en raison des allergènes déclarés."""
    if not user_allergens:
        return set()
    coffret = _get_coffret()
    allergen_map = coffret["allergen_map"]
    user_allergens_lower = {a.strip().lower() for a in user_allergens}
    blocked: set[str] = set()
    for ingredient_name, allergens in allergen_map.items():
        if {a.lower() for a in allergens} & user_allergens_lower:
            blocked.add(ingredient_name)
    return blocked


# ── Boosters ─────────────────────────────────────────────────────────

BOOSTERS = [
    {
        "name": "Floral",
        "keywords": ["fleur", "rose", "jasmin", "muguet", "floral", "flower",
                      "pétale", "bouquet", "pivoine", "iris", "ylang",
                      "néroli", "magnolia", "tubéreuse", "gardénia"],
    },
    {
        "name": "Ambre doux",
        "keywords": ["ambre", "vanille", "oriental", "chaud", "doux", "warm",
                      "amber", "gourmand", "caramel", "miel", "tonka",
                      "baume", "résine", "encens", "oud", "boisé"],
    },
    {
        "name": "Musc blanc sec",
        "keywords": ["musc", "propre", "frais", "clean", "musk", "coton",
                      "savon", "linge", "poudré", "aldéhyde", "blanc",
                      "minéral", "ozonic", "aquatique", "agrume",
                      "bergamote", "citron", "pamplemousse"],
    },
]


def _select_boosters(
    ingredients: dict[str, list[dict]],
    count: int = 1,
) -> list[dict]:
    """Sélectionne les meilleurs boosters par scoring de mots-clés sur les notes."""
    text_parts: list[str] = []
    for note_key in ("top_notes", "heart_notes", "base_notes"):
        for note in ingredients.get(note_key, []):
            for field in ("name", "family", "description"):
                val = note.get(field)
                if val:
                    text_parts.append(val.lower())
    combined = " ".join(text_parts)

    scored: list[tuple[dict, int]] = []
    for booster in BOOSTERS:
        score = sum(1 for kw in booster["keywords"] if kw in combined)
        scored.append((booster, score))

    scored.sort(key=lambda x: x[1], reverse=True)
    return [b for b, _ in scored[:count]]


# ── Calcul des quantités en ml ────────────────────────────────────────

def _compute_quantities(
    selected_notes: dict[str, list[dict]],
    booster: dict,
    formula_type: str,
    target_ml: int,
) -> dict:
    """Calcule les ml pour chaque note selon le type de formule et le volume cible."""
    config = _FORMULA_TYPE_CONFIGS[formula_type]["sizes"][target_ml]

    return {
        "target_ml": target_ml,
        "formula_type": formula_type,
        "top_notes": [
            {**note, "ml": config["top_ml"]}
            for note in selected_notes.get("top_notes", [])
        ],
        "heart_notes": [
            {**note, "ml": config["heart_ml"]}
            for note in selected_notes.get("heart_notes", [])
        ],
        "base_notes": [
            {**note, "ml": config["base_ml"]}
            for note in selected_notes.get("base_notes", [])
        ],
        "boosters": [{"name": booster["name"], "ml": config["booster_ml"]}],
    }


# ── Construction d'une formule ────────────────────────────────────────

def _build_formula(
    note_scores: dict[str, dict[str, float]],
    blocked_names: set[str],
    excluded_names: set[str],
    language: str,
    force_type: str | None = None,
) -> dict:
    """Construit une formule complète à partir des scores de notes.

    Retourne le dict formule enrichi d'une clé interne '_selected_en_names'
    (noms EN des notes utilisées) pour exclure ces notes de la formule suivante.
    """
    descriptions = PROFILE_DESCRIPTIONS_EN if language == "en" else PROFILE_DESCRIPTIONS
    translate_name = (lambda n: INGREDIENT_EN_TO_FR.get(n, n)) if language == "fr" else (lambda n: n)

    # 1. Sélection préliminaire max 3 par catégorie pour dériver le profil
    preliminary = _select_notes_by_score(
        note_scores,
        max_per_cat=3,
        excluded_names=excluded_names,
        blocked_names=blocked_names,
    )

    # 2. Dériver le profil dominant et le type de formule
    profile_name = _derive_profile_from_notes(preliminary)
    formula_type = force_type if force_type in _FORMULA_TYPE_CONFIGS else _classify_formula_type(profile_name)

    # 3. Ajuster la sélection au nombre de notes du type de formule
    type_counts = _FORMULA_TYPE_CONFIGS[formula_type]["note_counts"]
    selected_notes = {
        "top_notes": preliminary["top_notes"][: type_counts["top"]],
        "heart_notes": preliminary["heart_notes"][: type_counts["heart"]],
        "base_notes": preliminary["base_notes"][: type_counts["base"]],
    }

    # 4. Sélectionner le booster (1 seul)
    booster_list = _select_boosters(selected_notes, count=1)
    booster = booster_list[0] if booster_list else BOOSTERS[0]

    # 5. Collecter les noms EN avant traduction (pour l'exclusion formule suivante)
    selected_en_names = {n["name"] for v in selected_notes.values() for n in v}

    # 6. Traduire les noms d'ingrédients pour l'affichage
    translated = {
        key: [{**n, "name": translate_name(n["name"])} for n in notes]
        for key, notes in selected_notes.items()
    }

    # 7. Calculer les quantités pour les 3 formats
    sizes = {}
    for target_ml in (10, 30, 50):
        sizes[f"{target_ml}ml"] = _compute_quantities(
            translated, booster, formula_type, target_ml
        )

    return {
        "profile": profile_name,
        "formula_type": formula_type,
        "description": descriptions.get(profile_name, ""),
        "top_notes": [n["name"] for n in translated["top_notes"]],
        "heart_notes": [n["name"] for n in translated["heart_notes"]],
        "base_notes": [n["name"] for n in translated["base_notes"]],
        "details": translated,
        "sizes": sizes,
        "_selected_en_names": selected_en_names,  # clé interne, retirée avant stockage
    }


# ── Génération des formules ───────────────────────────────────────────

def generate_formulas(session_id: str, force_type: str | None = None) -> dict:
    """Génère 2 formules personnalisées pour une session.

    1. Récupère les réponses depuis Redis
    2. Score les notes directement depuis les réponses (note_scoring_mapping.json)
    3. Dérive le profil et le type de formule depuis les notes sélectionnées
    4. Retourne 2 formules avec les notes et les ml selon le type
    """
    session_data = session_store.get_session_answers(session_id)
    if not session_data or not session_data.get("answers"):
        return {"error": "Aucune réponse trouvée", "formulas": []}

    session_meta = session_store.get_session_meta(session_id)
    language = session_meta.get("language", "fr") if session_meta else "fr"

    profile = session_store.get_user_profile(session_id)
    has_allergies = profile.get("has_allergies", "non") if profile else "non"
    user_allergens_raw = profile.get("allergies", "") if profile else ""

    user_allergens = None
    if has_allergies == "oui" and user_allergens_raw:
        user_allergens = [
            a.strip()
            for a in user_allergens_raw.replace(",", ";").split(";")
            if a.strip()
        ]

    blocked_names = _get_blocked_ingredients(user_allergens)

    # Scorer toutes les notes depuis les réponses
    note_scores = _score_notes(session_data["answers"])

    # Générer 2 formules avec des notes différentes
    formulas = []
    excluded_names: set[str] = set()

    for _ in range(2):
        formula = _build_formula(note_scores, blocked_names, excluded_names, language, force_type)
        excluded_names |= formula.pop("_selected_en_names", set())
        formulas.append(formula)

    session_store.save_generated_formulas(session_id, formulas)
    return {"formulas": formulas}


# ── Sélection et personnalisation ─────────────────────────────────────

def change_selected_formula_type(session_id: str, formula_type: str) -> dict:
    """Régénère une seule formule avec un nouveau type et la sauvegarde directement comme sélectionnée."""
    if formula_type not in _FORMULA_TYPE_CONFIGS:
        return {"error": f"formula_type must be one of: {', '.join(_FORMULA_TYPE_CONFIGS)}"}

    session_data = session_store.get_session_answers(session_id)
    if not session_data or not session_data.get("answers"):
        return {"error": "Aucune réponse trouvée"}

    session_meta = session_store.get_session_meta(session_id)
    language = session_meta.get("language", "fr") if session_meta else "fr"

    profile = session_store.get_user_profile(session_id)
    has_allergies = profile.get("has_allergies", "non") if profile else "non"
    user_allergens_raw = profile.get("allergies", "") if profile else ""

    user_allergens = None
    if has_allergies == "oui" and user_allergens_raw:
        user_allergens = [
            a.strip()
            for a in user_allergens_raw.replace(",", ";").split(";")
            if a.strip()
        ]

    blocked_names = _get_blocked_ingredients(user_allergens)
    note_scores = _score_notes(session_data["answers"])

    formula = _build_formula(note_scores, blocked_names, set(), language, force_type=formula_type)
    formula.pop("_selected_en_names", None)

    session_store.save_selected_formula(session_id, formula)
    return {"formula": formula}


def select_formula(session_id: str, formula_index: int) -> dict:
    """Sélectionne une des 2 formules générées et la stocke dans Redis."""
    formulas = session_store.get_generated_formulas(session_id)
    if not formulas:
        return {"error": "No generated formulas found"}
    if formula_index not in (0, 1):
        return {"error": "formula_index must be 0 or 1"}
    if formula_index >= len(formulas):
        return {"error": "Invalid formula index"}

    selected = formulas[formula_index]
    session_store.save_selected_formula(session_id, selected)

    return {"formula": selected}


def get_available_ingredients(
    session_id: str,
    note_type: str,
) -> dict:
    """Retourne tous les ingrédients disponibles pour un type de note, filtré par allergènes."""
    if note_type not in ("top", "heart", "base"):
        return {"error": "note_type must be top, heart, or base"}

    profile = session_store.get_user_profile(session_id)
    has_allergies = profile.get("has_allergies", "non") if profile else "non"
    user_allergens_raw = profile.get("allergies", "") if profile else ""

    user_allergens = None
    if has_allergies == "oui" and user_allergens_raw:
        user_allergens = [
            a.strip()
            for a in user_allergens_raw.replace(",", ";").split(";")
            if a.strip()
        ]

    session_meta = session_store.get_session_meta(session_id)
    language = session_meta.get("language", "fr") if session_meta else "fr"
    translate_name = (
        (lambda name: INGREDIENT_EN_TO_FR.get(name, name)) if language == "fr"
        else (lambda name: name)
    )

    coffret = _get_coffret()
    blocked_ingredients = _get_blocked_ingredients(user_allergens)

    ingredients = []
    for ingredient in coffret["ingredients"]:
        if ingredient["note_type"] != note_type:
            continue
        if ingredient["name"] in blocked_ingredients:
            continue
        ingredients.append({
            "name": translate_name(ingredient["name"]),
            "family": ingredient["family"],
            "description": ingredient["description"],
        })

    return {"note_type": note_type, "ingredients": ingredients}


def replace_note(
    session_id: str,
    note_type: str,
    old_note: str,
    new_note: str,
) -> dict:
    """Remplace une note dans la formule sélectionnée et recalcule les ml."""
    if note_type not in ("top", "heart", "base"):
        return {"error": "note_type must be top, heart, or base"}

    selected = session_store.get_selected_formula(session_id)
    if not selected:
        return {"error": "No formula selected yet"}

    note_key = {"top": "top_notes", "heart": "heart_notes", "base": "base_notes"}[note_type]
    formula_type = selected.get("formula_type", "mix")

    session_meta = session_store.get_session_meta(session_id)
    language = session_meta.get("language", "fr") if session_meta else "fr"
    translate_name = (
        (lambda name: INGREDIENT_EN_TO_FR.get(name, name)) if language == "fr"
        else (lambda name: name)
    )

    coffret = _get_coffret()
    new_ingredient = None
    for ingredient in coffret["ingredients"]:
        if ingredient["note_type"] != note_type:
            continue
        translated = translate_name(ingredient["name"])
        if translated.lower() == new_note.lower() or ingredient["name"].lower() == new_note.lower():
            new_ingredient = {
                "name": translated,
                "family": ingredient["family"],
                "description": ingredient["description"],
                "position": ingredient["position"],
            }
            break

    if not new_ingredient:
        return {"error": f"Ingredient '{new_note}' not found in coffret for {note_type} notes"}

    details = selected.get("details", {})
    found = False
    for i, note in enumerate(details.get(note_key, [])):
        if note["name"].lower() == old_note.lower():
            details[note_key][i] = new_ingredient
            found = True
            break

    if not found:
        return {"error": f"Note '{old_note}' not found in current formula's {note_key}"}

    selected[note_key] = [n["name"] for n in details[note_key]]
    selected["details"] = details

    # Récupérer le booster existant et recalculer les ml
    booster_list = selected.get("sizes", {}).get("30ml", {}).get("boosters", [])
    booster = {"name": booster_list[0]["name"] if booster_list else BOOSTERS[0]["name"], "keywords": []}

    sizes = {}
    for target_ml in (10, 30, 50):
        sizes[f"{target_ml}ml"] = _compute_quantities(details, booster, formula_type, target_ml)

    selected["sizes"] = sizes
    session_store.save_selected_formula(session_id, selected)

    return {"formula": selected}
